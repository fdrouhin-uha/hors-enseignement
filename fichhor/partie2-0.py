from pandas import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import  get_column_letter


def lire_fichier_entree(nom_fichier, feuille):
    try:
        df = pd.read_excel(nom_fichier, sheet_name=feuille)
        return df
    except Exception as e:
        print("Une erreur s'est produite lors de la lecture du fichier d'entrée :", str(e))
        return None


def ajuster_largeur_colonnes(nom_fichier):
    try:
        wb = load_workbook(nom_fichier)
        ws = wb.active
        for column in ws.columns:
            max_length = 0
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
        wb.save(nom_fichier)
    except Exception as e:
        print("Une erreur s'est produite lors de l'ajustement de la largeur des colonnes :", str(e))


def ecrire_fichier_sortie(df_code, df_coef,nom_fichier):
    try:
        coef=[]
        retour=[]
        i=0
        total=0
        general = []
        generaltot = 0
        for index, row in df_coef.iterrows():
            temp=row["Nombre d'heures équivalent TD"]
            coef.append(temp)
        for y in range(len(coef)):
            general.append(0)
        for nom_colonne in df_code.columns[1:]:
            df_subset = df_code[['Référentiel', nom_colonne]]
            for index, row in df_subset.iterrows():
                ligne=row[nom_colonne]
                if isinstance(ligne, int):
                    retour.append(ligne*coef[i])
                    total += ligne*coef[i]
                    general[i] += ligne*coef[i]
                else:
                    retour.append(0)
                i+=1

            df_subset = df_code[['Référentiel']]
            df_subset = df_subset.assign(temp=retour)
            df_subset = df_subset.rename(columns={'temp': nom_colonne})
            df_subset = pd.concat([ df_subset, pd.DataFrame([{"Référentiel": "Total", nom_colonne: total}])], ignore_index=True)
            nom_fichier_sortie = nom_fichier.split('.')[0] + f"_{nom_colonne}.xlsx"
            df_subset.to_excel(nom_fichier_sortie, index=False, sheet_name='Feuil1')
            ajuster_largeur_colonnes(nom_fichier_sortie)
            print("Le fichier Excel de sortie a été créé avec succès :", nom_fichier_sortie)

            generaltot += total
            i=0
            total = 0
            retour=[]
            
        df_total = df_code[['Référentiel']]
        df_total = df_total.assign(total=general)
        df_total = pd.concat([df_total, pd.DataFrame([{"Référentiel": "Total", "Total": generaltot}])], ignore_index=True)
        df_total.to_excel("fichier_total.xlsx", index=False, sheet_name='Feuil1')
        ajuster_largeur_colonnes("fichier_total.xlsx")
        print("Les données ont été écrites avec succès dans les fichiers Excel de destination.")
    except Exception as e:
       print("Une erreur s'est produite lors de l'écriture des fichiers de sortie :", str(e))


fichier_coef = "D:/zCour/stage/premier_test/fichor_sortie.xlsx"
feuille_coef = "Sheet1"
df_coeficien = lire_fichier_entree(fichier_coef, feuille_coef)


fichier_ens = "D:/zCour/stage/premier_test/fichor_ens_sortie.xlsx" 
feuille_ens = "Sheet1"  
df_ens = lire_fichier_entree(fichier_ens, feuille_ens)


if df_ens is not None:
    fichier_horaire = "fichor_sortie.xlsx"
    ecrire_fichier_sortie(df_ens, df_coeficien, fichier_horaire)