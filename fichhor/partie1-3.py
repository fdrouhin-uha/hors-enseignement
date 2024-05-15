import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import  get_column_letter

def lire_fichier_entree(nom_fichier, feuille):
    try:
        df = pd.read_excel(nom_fichier, sheet_name=feuille)
        return df
    except Exception as e:
        print("Une erreur s'est produite lors de la lecture du fichier d'entrée :", str(e))
        return None

def ecrire_fichier_sortie(df_code, df_ens, nom_fichier, nom_fich_ens):
    try:
        data = []
        for index, row in df_code.iterrows():
            code_etape = row['Code Etape']
            intitule_code_etape = row['Intitulé code étape']
            referenciel = intitule_code_etape
            if isinstance(code_etape,str):
                if 'FA' in intitule_code_etape:
                    Apprenti_heures = row['Apprenti'] if pd.notnull(row['Apprenti']) else 0
                    sae_suivi_heures = row['SAE suivi'] if pd.notnull(row['SAE suivi']) else 0
                    sae_responsable_heures = row['SAE responsable'] if pd.notnull(row['SAE responsable']) else 0
                    data.append({'Référentiel': f"{referenciel} - Apprenti", 'Code étape': code_etape, 'Nombre d\'heures équivalent TD': Apprenti_heures, 'Remarque': 'heures par apprenti'})
                    data.append({'Référentiel': f"{referenciel} - SAE suivi", 'Code étape': code_etape, 'Nombre d\'heures équivalent TD': sae_suivi_heures, 'Remarque': 'heures de suivis de SAE'})
                    data.append({'Référentiel': f"{referenciel} - SAE responsable", 'Code étape': code_etape, 'Nombre d\'heures équivalent TD': sae_responsable_heures, 'Remarque': 'heures de responsabilités de SAE'})
                elif 'FI' in intitule_code_etape:
                    stage_heures = row['Stage'] if pd.notnull(row['Stage']) else 0
                    sae_suivi_heures = row['SAE suivi'] if pd.notnull(row['SAE suivi']) else 0
                    sae_responsable_heures = row['SAE responsable'] if pd.notnull(row['SAE responsable']) else 0
                    data.append({'Référentiel': f"{referenciel} - Stage", 'Code étape': code_etape, 'Nombre d\'heures équivalent TD': stage_heures, 'Remarque': 'heures par stagiaire'})
                    data.append({'Référentiel': f"{referenciel} - SAE suivi", 'Code étape': code_etape, 'Nombre d\'heures équivalent TD': sae_suivi_heures, 'Remarque': 'heures de suivis de SAE'})
                    data.append({'Référentiel': f"{referenciel} - SAE responsable", 'Code étape': code_etape, 'Nombre d\'heures équivalent TD': sae_responsable_heures, 'Remarque': 'heures de responsabilités de SAE'}) 
                elif 'FI' not in intitule_code_etape and 'FA' not in intitule_code_etape:
                    stage_heures = row['Stage'] if pd.notnull(row['Stage']) else 0
                    sae_suivi_heures = row['SAE suivi'] if pd.notnull(row['SAE suivi']) else 0
                    sae_responsable_heures = row['SAE responsable'] if pd.notnull(row['SAE responsable']) else 0
                    Apprenti_heures = row['Apprenti'] if pd.notnull(row['Apprenti']) else 0
                    data.append({'Référentiel': f"{referenciel} - Apprenti", 'Code étape': code_etape, 'Nombre d\'heures équivalent TD': Apprenti_heures, 'Remarque': 'heures par apprenti'})
                    data.append({'Référentiel': f"{referenciel} - Stage", 'Code étape': code_etape, 'Nombre d\'heures équivalent TD': stage_heures, 'Remarque': 'heures par stagiaire'})
                    data.append({'Référentiel': f"{referenciel} - SAE suivi", 'Code étape': code_etape, 'Nombre d\'heures équivalent TD': sae_suivi_heures, 'Remarque': 'heures de suivis de SAE'})
                    data.append({'Référentiel': f"{referenciel} - SAE responsable", 'Code étape': code_etape, 'Nombre d\'heures équivalent TD': sae_responsable_heures, 'Remarque': 'heures de responsabilités de SAE'})         
        data.extend([
            {'Référentiel': "Loi ORE : mettre 1 si l'enseignant a étudié des dossiers, 0 sinon", 'Code étape': " ", 'Nombre d\'heures équivalent TD': 0.5, 'Remarque': "h pour relecture du dossier"}
            ])
        
        nouveau_df = pd.DataFrame(data)
        nouveau_df.to_excel(nom_fichier, index=False, sheet_name='Feuil1')

        ajuster_largeur_colonnes(nom_fichier)

        data_ens=cree_ens(nouveau_df)

        nouveau_df_ens = pd.DataFrame(data_ens)
        nouveau_df_ens = nouveau_df_ens.fillna('')
        nouveau_df_ens = nouveau_df_ens.transpose()
        nouveau_df_ens.reset_index(inplace=True)
        nouveau_df_ens = nouveau_df_ens.rename(columns={'index': 'Référentiel'})
        nouveau_df_ens.to_excel(nom_fich_ens, index=False, sheet_name='Feuil1')

        ajuster_largeur_colonnes(nom_fich_ens)
        
        print("Les données ont été écrites avec succès dans le fichier Excel de destination :", nom_fichier)
        print("La fiche d'enseignant a été écrite dans le fichier de destination", nom_fich_ens)
    except Exception as e:
        print("Une erreur s'est produite lors de l'écriture du fichier de sortie :", str(e))

def cree_ens(nouveau_df):
    try:
        data = {}
        for index, row in nouveau_df.iterrows():
            referenciel = row['Référentiel']
            if isinstance(referenciel,str):
                if referenciel not in data:
                    data[referenciel] = {}
                for index, row_ens in df_ens.iterrows():
                    if isinstance(row_ens['NOM'], str) or isinstance(row_ens['Prénom'], str):
                        enseignant = row_ens['NOM'] + ' ' + row_ens['Prénom']
                    if isinstance(enseignant, str):
                        if enseignant not in data[referenciel]:
                            data[referenciel][enseignant] = ""
        return data
    except Exception as e:
        print("Une erreur s'est produite lors de la creation du fichier enseignants", str(e))

def ajuster_largeur_colonnes(nom_fichier):
    try:
        wb = load_workbook(nom_fichier)
        ws = wb.active
        for column in ws.columns:
            max_length = 0
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
        wb.save(nom_fichier)
    except Exception as e:
        print("Une erreur s'est produite lors de l'ajustement de la largeur des colonnes :", str(e))


fichier_ens = "D:/zCour/stage/premier_test/fichhor/list_ens.xlsx"
feuille_entree_ens = "Feuil1"
df_ens = lire_fichier_entree(fichier_ens, feuille_entree_ens)


nom_fichier_entree = "D:/zCour/stage/premier_test/fichhor/Codes étapes TC.xlsx" 
feuille_entree = "Feuil1"  
df_entree = lire_fichier_entree(nom_fichier_entree, feuille_entree)


if df_entree is not None:
    nom_fichier_horaire = "fichor_sortie.xlsx"
    nom_fichier_horaire_ens = "fichor_ens_sortie.xlsx"
    ecrire_fichier_sortie(df_entree, df_ens, nom_fichier_horaire, nom_fichier_horaire_ens)