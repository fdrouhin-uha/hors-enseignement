import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import  get_column_letter

def sortie_groupe(df, nom_fichier):
    try:
        heures_agreg = {
            "TD": {},  # Dictionnaire pour les heures de TD
            "TP": {},  # Dictionnaire pour les heures de TP
            "CM": {}   # Dictionnaire pour les heures de CM
        }
        
        combinaisons = set()  # Ensemble pour stocker les combinaisons groupe-enseignant déjà ajoutées
        
        for index, row in df.iterrows():
            groupe_all = row["Groupes d'étudiants imposés (noms)"]
            ens_all = row["Enseignants"]
            type_enseignement = row["Type"]
            heures = row["Durée (h)"]
            
            # Diviser les groupes en utilisant ", " et garder seulement le premier
            if isinstance(groupe_all, str):
                groupe = groupe_all.split(", ")[0]
            else:
                groupe = str(groupe_all)
            
            # Diviser les enseignants sur ", " pour obtenir une liste d'enseignants
            if isinstance(ens_all, str):
                enseignants = ens_all.split(', ')
            else:
                enseignants = [str(ens_all)]
            
            # Convertir l'heure en nombre
            heures_nombre = convertir_heure_en_nombre(heures)
            
            # Si le type d'enseignement est "TD", "TP" ou "CM" et s'il y a un enseignant
            if isinstance(ens_all, str) and type_enseignement in ["TD", "TP", "CM"]:
                for enseignant in enseignants:
                    # Créer une clé basée sur [groupe, enseignant]
                    key = (groupe, enseignant)
                    
                    # Vérifier si la combinaison groupe-enseignant n'a pas déjà été ajoutée
                    if key not in combinaisons:
                        # Ajouter la combinaison à l'ensemble
                        combinaisons.add(key)
                        
                        # Ajouter les heures à la valeur existante ou initialiser à 0 si elle n'existe pas encore
                        heures_agreg[type_enseignement][key] = heures_agreg[type_enseignement].get(key, 0) + heures_nombre
        
        # Créer une liste pour stocker les données agrégées
        retour = []
        for type_enseignement, heures_dict in heures_agreg.items():
            for key, heures in heures_dict.items():
                groupe, enseignant = key
                heures_totales, minutes = divmod(heures, 60)
                heure_formattee = '{:02d}h{:02d}'.format(heures_totales, minutes)
                retour.append({'Groupe': groupe, 'Enseignant': enseignant, 'Type': type_enseignement, 'Heure': heure_formattee})
        
        # Créer un DataFrame à partir de la liste
        nouveau_df = pd.DataFrame(retour)
        
        # Sauvegarder le DataFrame dans un fichier Excel
        nouveau_df.to_excel(nom_fichier, index=False)   
        
        wb = load_workbook(nom_fichier)
        ws = wb.active
        for column in ws.columns:
            max_length = 0
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

        wb.save(nom_fichier)
        print("Les données ont été écrites avec succès dans le fichier Excel de destination :", nom_fichier)
    except Exception as e:
        print("Une erreur s'est produite lors de l'écriture du fichier de sortie :", str(e))

def convertir_heure_en_nombre(heure_str):
    # Séparer l'heure et les minutes en fonction de 'h'
    heures, minutes = heure_str.split('h')
    # Convertir les heures et les minutes en nombres entiers
    heures = int(heures)
    minutes = int(minutes)
    # Calculer le total en minutes
    total_minutes = heures * 60 + minutes
    return total_minutes

def lire_fichier_entree(nom_fichier, feuille):
    try:
        return pd.read_excel(nom_fichier, feuille)
    except Exception as e:
        print("Une erreur s'est produite lors de la lecture du fichier d'entrée :", str(e))
        return None

nom_fichier_entree = "D:\zCour\stage\premier_test\hse\HSE.xlsx"
feuille_entree = "Feuil1"  
df_entree = lire_fichier_entree(nom_fichier_entree, feuille_entree)

if df_entree is not None:
    nom_fichier_sortie = "hse_groupe_enseignant_sortie.xlsx"
    sortie_groupe(df_entree, nom_fichier_sortie)
