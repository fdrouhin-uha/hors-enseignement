

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def sortie_groupe(df,df_correspondance, nom_fichier):
    try:
        heures_agreg = {
            "TD": {},  # Dictionnaire pour les heures de TD
            "TP": {},  # Dictionnaire pour les heures de TP
            "CM": {}   # Dictionnaire pour les heures de CM
        }
        
        combinaisons = set()  # Ensemble pour stocker les combinaisons groupe-enseignant-type déjà ajoutées
        
        for index, row in df.iterrows():
            groupe_all = row["Groupes d'étudiants imposés (noms)"]
            ens_all = row["Enseignants"]
            type_enseignement = row["Type"]
            heures = row["Durée (h)"]
            

            
            # Diviser les groupes en utilisant ", " et garder seulement le premier
            if isinstance(groupe_all, str):
                groupe = groupe_all.split(", ")[0]
            else:
                groupe = str(groupe_all)
            
            # Diviser les enseignants sur ", " pour obtenir une liste d'enseignants
            if isinstance(ens_all, str):
                    enseignants = [enseignant.strip() for enseignant in str(ens_all).split(",")]
                    
            else:
                enseignants = [str(ens_all)]
            
            # Convertir l'heure en nombre
            heures_nombre = convertir_heure_en_nombre(heures)
            
            # Si le type d'enseignement est "TD", "TP" ou "CM" et s'il y a un enseignant
            if isinstance(ens_all, str) and type_enseignement in ["TD", "TP", "CM"]:
                for enseignant in enseignants:
                    # Créer une clé basée sur [groupe, enseignant, type_enseignement]
                    for index, row in df_correspondance.iterrows():
                        groupe_cores=row["Groupes"]
                        if groupe == groupe_cores:
                            groupe = row["Correspondance"]
                            key = (groupe, enseignant, type_enseignement)
                    
                    # Vérifier si la combinaison groupe-enseignant-type n'a pas déjà été ajoutée
                    if key not in combinaisons:
                        # Ajouter la combinaison à l'ensemble
                        combinaisons.add(key)
                        heures_agreg[type_enseignement][key] = heures_agreg[type_enseignement].get(key, 0) + heures_nombre
                    else :
                        # Ajouter les heures à la valeur existante ou initialiser à 0 si elle n'existe pas encore
                        heures_agreg[type_enseignement][key] = heures_agreg[type_enseignement].get(key, 0) + heures_nombre
        
        # Créer une liste pour stocker les données agrégées
        retour = []
        for (groupe, enseignant, _), heures in heures_agreg["TD"].items():
            if (groupe, enseignant) not in combinaisons:
                heure_td = heures_agreg["TD"].get((groupe, enseignant, "TD"), 0) / 60  # Convertir en heures décimales
                heure_tp = heures_agreg["TP"].get((groupe, enseignant, "TP"), 0) / 60  # Convertir en heures décimales
                heure_cm = heures_agreg["CM"].get((groupe, enseignant, "CM"), 0) / 60  # Convertir en heures décimales
                
                retour.append({'Groupe': groupe, 'Enseignant': enseignant, 'Heure_TD': heure_td, 'Heure_TP': heure_tp, 'Heure_CM': heure_cm}) 
                combinaisons.add((groupe, enseignant))  # Ajouter la combinaison au dictionnaire temporaire
        
        for (groupe, enseignant, _), heures in heures_agreg["TP"].items():
            if (groupe, enseignant) not in combinaisons:
                heure_td = heures_agreg["TD"].get((groupe, enseignant, "TD"), 0) / 60  # Convertir en heures décimales
                heure_tp = heures_agreg["TP"].get((groupe, enseignant, "TP"), 0) / 60  # Convertir en heures décimales
                heure_cm = heures_agreg["CM"].get((groupe, enseignant, "CM"), 0) / 60  # Convertir en heures décimales
                
                retour.append({'Groupe': groupe, 'Enseignant': enseignant, 'Heure_TD': heure_td, 'Heure_TP': heure_tp, 'Heure_CM': heure_cm}) 
                combinaisons.add((groupe, enseignant))  # Ajouter la combinaison au dictionnaire temporaire

        for (groupe, enseignant, _), heures in heures_agreg["CM"].items():
            if (groupe, enseignant) not in combinaisons:
                heure_td = heures_agreg["TD"].get((groupe, enseignant, "TD"), 0) / 60  # Convertir en heures décimales
                heure_tp = heures_agreg["TP"].get((groupe, enseignant, "TP"), 0) / 60  # Convertir en heures décimales
                heure_cm = heures_agreg["CM"].get((groupe, enseignant, "CM"), 0) / 60  # Convertir en heures décimales
                
                retour.append({'Groupe': groupe, 'Enseignant': enseignant, 'Heure_TD': heure_td, 'Heure_TP': heure_tp, 'Heure_CM': heure_cm}) 
                combinaisons.add((groupe, enseignant)) 
        
        # Créer un DataFrame à partir de la liste
        nouveau_df = pd.DataFrame(retour)
        
        # Trier les données par groupe, enseignant, et type
        nouveau_df = nouveau_df.sort_values(by=['Groupe', 'Enseignant'])
        
        # Sauvegarder le DataFrame dans un fichier Excel
        nouveau_df.to_excel(nom_fichier, index=False)
        
        # Ajuster la largeur des colonnes dans le fichier Excel
        ajuster_largeur_colonnes(nom_fichier)
        
        print("Les données ont été écrites avec succès dans le fichier Excel de destination :", nom_fichier)
    except Exception as e:
        print("Une erreur s'est produite lors de l'écriture du fichier de sortie :", str(e))

def ajuster_largeur_colonnes(nom_fichier):
    try:
        wb = load_workbook(nom_fichier)
        ws = wb.active
        for column in ws.columns:
            max_length = 0
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
        wb.save(nom_fichier)  # N'oubliez pas de sauvegarder le classeur Excel après l'ajustement
    except Exception as e:
        print("Une erreur s'est produite lors de l'ajustement de la largeur des colonnes :", str(e))


def convertir_heure_en_nombre(heure_str):
    # Séparer l'heure et les minutes en fonction de 'h'
    heures, minutes = heure_str.split('h')
    # Convertir les heures et les minutes en nombres entiers
    heures = int(heures)
    minutes = int(minutes)
    # Calculer le total en minutes
    total_minutes = heures * 60 + minutes
    return total_minutes

def lire_fichier_entree(nom_fichier, feuille):
    try:
        return pd.read_excel(nom_fichier, feuille)
    except Exception as e:
        print("Une erreur s'est produite lors de la lecture du fichier d'entrée :", str(e))
        return None

nom_fichier_entree = "D:\zCour\stage\premier_test\hse\HSE.xlsx"
feuille_entree = "Feuil1"  
df_entree = lire_fichier_entree(nom_fichier_entree, feuille_entree)

nom_fichier_correspondance = "D:\zCour\stage\Copie_de_hse_groupe_sortie.xlsx"
feuille_entree_cor = "Sheet1"
df_correspondance = lire_fichier_entree(nom_fichier_correspondance, feuille_entree_cor)

if df_entree is not None:
    nom_fichier_sortie = "hse_groupe_enseignant_sortie_2.xlsx"
    sortie_groupe(df_entree,df_correspondance, nom_fichier_sortie)
