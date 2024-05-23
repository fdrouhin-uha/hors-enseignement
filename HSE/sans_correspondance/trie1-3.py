

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def sortie_groupe(df, nom_fichier):
    try:
        heures_agreg = {"TD": {}, "TP": {}, "CM": {}}
        
        combinaisons = set()
        
        for index, row in df.iterrows():
            groupe_all = row["Groupes d'étudiants imposés (noms)"]
            ens_all = row["Enseignants"]
            type_enseignement = row["Type"]
            heures = row["Durée (h)"]
            
            if isinstance(groupe_all, str):
                groupe = groupe_all.split(", ")[0]
            else:
                groupe = str(groupe_all)
            if isinstance(ens_all, str):
                enseignants = ens_all.split(', ')
            else:
                enseignants = [str(ens_all)]
            
            heures_nombre = convertir_heure_en_nombre(heures)

            if isinstance(ens_all, str) and type_enseignement in ["TD", "TP", "CM"]:
                for enseignant in enseignants:
                    key = (groupe, enseignant, type_enseignement)
                    if key not in combinaisons:
                        combinaisons.add(key)
                        heures_agreg[type_enseignement][key] = heures_agreg[type_enseignement].get(key, 0) + heures_nombre
                    else : 
                        heures_agreg[type_enseignement][key] = heures_agreg[type_enseignement].get(key, 0) + heures_nombre
        
        retour = []
        for (groupe, enseignant, _), heures in heures_agreg["TD"].items():
            if (groupe, enseignant) not in combinaisons:
                heure_td = heures_agreg["TD"].get((groupe, enseignant, "TD"), 0) / 60  # Convertir en heures décimales
                heure_tp = heures_agreg["TP"].get((groupe, enseignant, "TP"), 0) / 60  # Convertir en heures décimales
                heure_cm = heures_agreg["CM"].get((groupe, enseignant, "CM"), 0) / 60  # Convertir en heures décimales
                
                retour.append({'Groupe': groupe, 'Enseignant': enseignant, 'Heure_TD': heure_td, 'Heure_TP': heure_tp, 'Heure_CM': heure_cm}) 
                combinaisons.add((groupe, enseignant))  # Ajouter la combinaison au dictionnaire temporaire
        
        for (groupe, enseignant, _), heures in heures_agreg["TP"].items():
            if (groupe, enseignant) not in combinaisons:
                heure_td = heures_agreg["TD"].get((groupe, enseignant, "TD"), 0) / 60  # Convertir en heures décimales
                heure_tp = heures_agreg["TP"].get((groupe, enseignant, "TP"), 0) / 60  # Convertir en heures décimales
                heure_cm = heures_agreg["CM"].get((groupe, enseignant, "CM"), 0) / 60  # Convertir en heures décimales
                
                retour.append({'Groupe': groupe, 'Enseignant': enseignant, 'Heure_TD': heure_td, 'Heure_TP': heure_tp, 'Heure_CM': heure_cm}) 
                combinaisons.add((groupe, enseignant))  # Ajouter la combinaison au dictionnaire temporaire

        for (groupe, enseignant, _), heures in heures_agreg["CM"].items():
            if (groupe, enseignant) not in combinaisons:
                heure_td = heures_agreg["TD"].get((groupe, enseignant, "TD"), 0) / 60  # Convertir en heures décimales
                heure_tp = heures_agreg["TP"].get((groupe, enseignant, "TP"), 0) / 60  # Convertir en heures décimales
                heure_cm = heures_agreg["CM"].get((groupe, enseignant, "CM"), 0) / 60  # Convertir en heures décimales
                
                retour.append({'Groupe': groupe, 'Enseignant': enseignant, 'Heure_TD': heure_td, 'Heure_TP': heure_tp, 'Heure_CM': heure_cm}) 
                combinaisons.add((groupe, enseignant)) 
        
        nouveau_df = pd.DataFrame(retour)
        nouveau_df = nouveau_df.sort_values(by=['Groupe', 'Enseignant'])
        nouveau_df.to_excel(nom_fichier, index=False)
        ajuster_largeur_colonnes(nom_fichier)
        print("Les données ont été écrites avec succès dans le fichier Excel de destination :", nom_fichier)
    except Exception as e:
        print("Une erreur s'est produite lors de l'écriture du fichier de sortie :", str(e))

def ajuster_largeur_colonnes(nom_fichier):
    try:
        wb = load_workbook(nom_fichier)
        ws = wb.active
        for column in ws.columns:
            max_length = 0
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
        wb.save(nom_fichier)
    except Exception as e:
        print("Une erreur s'est produite lors de l'ajustement de la largeur des colonnes :", str(e))


def convertir_heure_en_nombre(heure_str):
    # Séparer l'heure et les minutes en fonction de 'h'
    heures, minutes = heure_str.split('h')
    # Convertir les heures et les minutes en nombres entiers
    heures = int(heures)
    minutes = int(minutes)
    # Calculer le total en minutes
    total_minutes = heures * 60 + minutes
    return total_minutes

def lire_fichier_entree(nom_fichier, feuille):
    try:
        return pd.read_excel(nom_fichier, feuille)
    except Exception as e:
        print("Une erreur s'est produite lors de la lecture du fichier d'entrée :", str(e))
        return None

nom_fichier_entree = "D:\zCour\stage\premier_test\hse\HSE.xlsx"
feuille_entree = "Feuil1"  
df_entree = lire_fichier_entree(nom_fichier_entree, feuille_entree)

if df_entree is not None:
    nom_fichier_sortie = "hse_groupe_enseignant_sortie.xlsx"
    sortie_groupe(df_entree, nom_fichier_sortie)
